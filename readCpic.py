import json
import os
import requests
import urllib
from openpyxl import Workbook #2.2.3
from openpyxl import load_workbook
import re
import itertools
from collections import OrderedDict
from zipfile import ZipFile

__METADATA__ = {
    "src_name": 'PharmGKB',
    "src_url": 'https://www.pharmgkb.org',
    "version": '0.1',
    "field": "pharmgkb"
}

translationtableUrls = {
'CYP2C19':'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1445556789&dlCls=HaplotypeSet&dlId=PA166128323&dlName=CPIC%20CYP2C19%20Haplotype%20Set',
'CYP2D6':'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1446767521&dlCls=HaplotypeSet&dlId=PA165980499&dlName=CYP2D6%20Cytochrome%20P450%20Nomenclature%20DB%20Haplotype%20Set',
'DPYD':'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1183805365&dlCls=HaplotypeSet&dlId=PA165980513&dlName=Haplotype%20Set%20PA165980513%20for%20DPYD',
'TMPT':'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1445164739&dlCls=HaplotypeSet&dlId=PA166128346&dlName=CPIC%20TMPT%20Haplotype%20Set',
'UGT1A1':'https://www.pharmgkb.org/download.do?objCls=SubmissionEvent&objId=1444880096&dlCls=HaplotypeSet&dlId=PA166115840&dlName=Haplotypes%20for%20UGT1A1%20(UGT%20Alleles%20Nomenclature%20page)'
}

jsonfilesUrl = 'https://www.pharmgkb.org/download.do?objId=dosingGuidelines.json.zip&dlCls=common'

def getGeneSymbolName(rsid):
    ''' Converts RsID to geneSymbolName through myvariant.info in iterations from 3 different annotation resources
    '''
    geneNameFromMyVariant = ''
    myvariantRsidRequest = requests.get('http://myvariant.info/v1/query?q='+rsid)
    if myvariantRsidRequest.status_code == requests.codes.ok:
        if bool(re.search('[r][s]\d+',rsid)):
            print 'http://myvariant.info/v1/query?q='+rsid
            commit_data = myvariantRsidRequest.json()
            try:
                print 'This is the hgvs id from myvariant.info:' ,commit_data['hits'][0]['_id']
            except IndexError:
                print 'invalid rs id ie no data found on myvariant.info'
                return
            print 'Searching for Genename...'
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['dbsnp']['gene']['symbol']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['dbsnp']['gene']['symbol']"""
            except TypeError, e:
                print """TypeError: Multiple genes found in ['hits'][0]['dbsnp']['gene']['symbol']""",e
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['dbsnp']['gene']['symbol']
            if geneNameFromMyVariant!='':
                return geneNameFromMyVariant
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['snpeff']['ann'][0]['gene_name']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['snpeff']['ann'][0]['gene_name']"""
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['snpeff']['ann'][0]['gene_name']
            if geneNameFromMyVariant!='':
                return geneNameFromMyVariant
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['dbnsfp']['genename']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['dbnsfp']['genename']"""
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['dbnsfp']['genename']
            if geneNameFromMyVariant!='':
                return geneNameFromMyVariant
            try:
                geneNameFromMyVariant=commit_data['hits'][0]['wellderly']['gene']
            except KeyError:
                pass
                # print """KeyError: Key not found in ['hits'][0]['dbnsfp']['genename']"""
            else:
                print 'Genename found! This is the gene from myvariant.info:' ,commit_data['hits'][0]['wellderly']['gene']
            if geneNameFromMyVariant!='' and not None and not type(geneNameFromMyVariant) is list:
                return geneNameFromMyVariant
            else:
                print 'genename not found on myvariant.info'
                raise ValueError('genename not found on myvariant.info or other error in genename search')
        else:
            print 'rsid malformed: '+rsid
            raise ValueError('rsid malformed: '+rsid)

    else:
        print '\nstatus_code at myvariant.info not ok!:',myvariantRsidRequest.headers['content-type']

def fillRsidList(translationTablePerGeneFileList):
    # print translationTablePerGeneFileList
    rsidList=[]
    for translationTablePerGene in translationTablePerGeneFileList:
        if translationTablePerGene.endswith('.xlsx'): #and translationTablePerGene.startswith(geneSymbolName):
            try:
                translationTablePerGeneWorkbook = load_workbook(translationTablePerGene,read_only=True)
            except ValueError:
                break
            worksheetTranslationTablePerGene = translationTablePerGeneWorkbook.active
            for row in worksheetTranslationTablePerGene.rows:
                for cell in row:
                    if isinstance(cell.value,unicode): #dont parse datetime objects
                        if str(cell.value.encode('utf8','ignore')).strip().startswith('rs'): #we are searching for all rs id #encoding due to unicode characters, str(unicode) gives unicodeencdoerror
                            rsidList.append(str(cell.value.encode('utf8','ignore')).strip())
    print 'number of rs#s found: ',len(rsidList)
    return rsidList

def load_data():
    urlopener = urllib.URLopener()
    translationTablePerGeneFileList = []
    for genename,url in translationtableUrls.iteritems():
        urlopener.retrieve(url,genename+'.xlsx')
        print 'downloading: ',genename+'.xlsx',' from: ',url[:50],' ...'
        translationTablePerGeneFileList.append(genename+'.xlsx')

    urlopener.retrieve(jsonfilesUrl,'dosingGuidelines_json.zip')
    dosingGuidelines_json_zipfile=ZipFile(open('dosingGuidelines_json.zip','rb'))

    pharmgkbJsonFileList=[]
    if not os.path.exists(os.getcwd()+'/dosingGuidelines_json'):
        os.makedirs(os.getcwd()+'/dosingGuidelines_json')
    for name in dosingGuidelines_json_zipfile.namelist():
        dosingGuidelines_json_zipfile.extract(name,os.getcwd()+'/dosingGuidelines_json')
        pharmgkbJsonFileList.append(os.getcwd()+'/dosingGuidelines_json/'+name)
    print 'downloading:  json files  from Pharmgkb.  ','Downloaded ',len(dosingGuidelines_json_zipfile.namelist()),' json files from ',jsonfilesUrl[:40],' ...'
    # print pharmgkbJsonFileList
    # fileOut = open('/Users/admin/Dropbox/Privat/00_Masterthesis/MITTELasdf.txt','w')

    rsidList = fillRsidList(translationTablePerGeneFileList)
    for rs in rsidList:
    print json.dumps(getDosingGuidelineFromRsid(rs,translationTablePerGeneFileList,pharmgkbJsonFileList), indent=4)
    # fileOut.write(json.dumps(getDosingGuidelineFromRsid('rs4244285',translationTablePerGeneFileList,pharmgkbJsonFileList), indent=4, sort_keys=True))
    # print json.dumps(getDosingGuidelineFromRsid('rs1801265'), indent=4)
    # fileOut.close()
    # getDosingGuidelineFromRsid('rs1801265')

def getHaplotypesFromTranslationtable(rsid,translationTablePerGeneFileList):
    for translationTablePerGene in translationTablePerGeneFileList:
        if translationTablePerGene.endswith('.xlsx') and not translationTablePerGene.startswith('~'):
            haplottypeListTemp=[]
            translationTablePerGeneWorkbook = load_workbook(translationTablePerGene,read_only=True)
            worksheetTranslationTablePerGene = translationTablePerGeneWorkbook.active
            coordinatesOfRsid = ''
            for row in worksheetTranslationTablePerGene.rows:
                for cell in row:
                    if isinstance(cell.value,unicode): #dont parse datetime objects
                        if str(cell.value.encode('utf8','ignore')).strip()==rsid: #encoding due to unicode characters, str(unicode) gives unicodeencdoerror
                            coordinatesOfRsid = cell.coordinate
            letterOfRsIdCell = ''
            if coordinatesOfRsid!='':
                letterOfRsIdCell = re.search('[A-Z]{1,2}', coordinatesOfRsid).group() #gives the letter of the coordinate
                rowCount = worksheetTranslationTablePerGene.get_highest_row()
                if not letterOfRsIdCell=='':
                    for i in range (1,rowCount+1):
                        try:
                            if worksheetTranslationTablePerGene[letterOfRsIdCell+str(i)].value: #take only non-empty cells
                                 if bool(re.search('\*\d',str(worksheetTranslationTablePerGene['B'+str(i)].value))): # pattern is star plus a digit then stop because we only want the basic star allels. We search in the B column because it contains the star alleles
                                    haplottypeListTemp.append(worksheetTranslationTablePerGene['B'+str(i)].value)
                        except IndexError, e:
                            print e
                        except:
                            pass
                print 'star alleles list:',haplottypeListTemp
                return haplottypeListTemp


def getDosingGuidelineFromRsid(rsid,translationTablePerGeneFileList,pharmgkbJsonFileList):
    try:
        geneSymbolName = getGeneSymbolName(rsid)
    except ValueError:
        return
    if geneSymbolName == None:
        return
    haplottypeListComplete = getHaplotypesFromTranslationtable(rsid,translationTablePerGeneFileList)
    if haplottypeListComplete==None:
        return
    if '*1' in haplottypeListComplete:
        haplottypeListComplete.remove('*1')

    print 'Searching for Dosing Guidelines for all ',len(haplottypeListComplete), 'star alleles.'

    jsonSnp = OrderedDict()
    jsonSnp['_id'] = rsid
    jsonSnp['pharmgkb'] = {}
    jsonSnp['pharmgkb']['rsid'] = rsid
    jsonSnp['pharmgkb']['gene'] = geneSymbolName
    jsonSnp['pharmgkb']['haplotypes'] = []
    jsonSnp['pharmgkb']['drugrecommendations'] = []
    for starAllele in haplottypeListComplete:
        jsonSnp['pharmgkb']['haplotypes'].append(geneSymbolName+starAllele)
        # if len(starAllelesListTwoBasic)>=2:
            # print 'Searching for Dosing Guidelines for star alleles:',starAllelesListTwoBasic[0],starAllelesListTwoBasic[1]
        for dosingGuidelinesJsonFile in pharmgkbJsonFileList:
            if geneSymbolName in dosingGuidelinesJsonFile:                        
                '''  takes a json file and searches for the two given star alleles and tries to print all found dosing guidelines
                '''
                with open(dosingGuidelinesJsonFile) as data_file:
                    parsedJsonFile = json.loads(data_file.read())
                if 'groups' in parsedJsonFile:
                    for groupsLoop in parsedJsonFile['groups']:
                        if 'genotypes' in groupsLoop:
                            levelOfEvidence=''
                            rec=''
                            drug = ''
                            for diplotypeLoop in groupsLoop['genotypes']:
                                for diplotypeLoopSplitUp in diplotypeLoop.split(';'):
                                    if geneSymbolName in diplotypeLoopSplitUp and starAllele in re.compile(r'\*\d+\b').findall(diplotypeLoopSplitUp):
                                        drug = parsedJsonFile['relatedDrugs'][0]['name']
                                        levelOfEvidence = groupsLoop['strength']['term']
                                        for annotationsLoop in groupsLoop['annotations']:
                                            if annotationsLoop['type']['term']=='Recommendations':
                                                rec = annotationsLoop['text']
                                            if (drug and rec and levelOfEvidence):
                                                jsonSnp['pharmgkb']['drugrecommendations'].append({'haplotypes':diplotypeLoop,'drug':drug,'recommendation':rec,'level_of_evidence':levelOfEvidence})
    if jsonSnp['pharmgkb']['drugrecommendations']==[]:
        return {}
    else:
        return jsonSnp
